from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl import load_workbook


class XlWriter:  # 写入xlsx
    def __init__(self, save_name, ws_name):
        """save_name文件名，ws_name表名"""
        self.save_name = save_name
        self.ws_name = ws_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = ws_name

    def append_row(self, row):
        """增加行
        :param row: row为一行元素的列表
        """
        self.ws.append(row)

    def select_sheet(self, title):
        """切换表"""
        if title not in self.wb.sheetnames:
            print(f'当前表格不含{title}页')
            return None
        self.ws = self.wb[title]

    def save(self):
        """保存文件"""
        # for ws in self.wb.worksheets:
        #     self.set_border(ws)
        self.wb.save(self.save_name)
        print(f'save to {self.save_name}')

    def new_sheet(self, title, row):
        """创建新的表，并切换到这个表
        :param title: 表名
        :param row: 第一行属性名
        """
        ws = self.wb.create_sheet(title=title)
        ws.append(row)
        self.ws = self.wb[title]

    def edit_cell(self, adderss, value):
        """修改单个单元格里内容
        :param adderss: excel里的坐标
        :param value: 修改的值
        """
        self.ws[adderss].value = value

    def set_border(self, ws):
        """设置表格格式"""
        for row in ws.rows:
            for cell in row:
                cell.font = Font(name=u'微软雅黑', size=11)
                thin = Side(border_style='thin', color='030303')
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


class XlReader:  # 读取xlsx
    def __init__(self, file):
        self.wb = load_workbook(file)

    def read(self, title, skip=[2],header_line=1):
        """
        :param title: 表名
        :param skip: 范围列表，前两个值为查询excel的行号范围，默认全部
        :return: 返回从第二行开始每行由header和value构成的字典的列表
        [{'head1': 'A2', 'head2': 'B2'}]
        """
        if title not in self.wb.sheetnames:
            print(f'当前表格不含{title}页')
            return False
        ws = self.wb[title]
        rules = []
        headers = None
        if len(skip) == 1: skip.append(len(list(ws.values)))
        if skip[0] < header_line+1: skip[0] = header_line+1
        for i, row in enumerate(ws.values):
            if i == header_line-1:
                headers = row
                continue
            elif skip[0] - 1 <= i <= skip[1] - 1:
                rule = {}
                for v in range(len(row)):
                    rule[headers[v]] = row[v]
                rules.append(rule)
        return rules

    def read_by_header(self, title, header, skip=[2],header_line=1):
        """
        :param title: 表名
        :param skip: 范围列表，前两个值为查询excel的行号范围，默认全部
        :return: 返回从第二行开始每行由指定header和value构成的字典的列表
        [{'head1': 'A2'}, {'head1': 'A3'}]
        """
        if title not in self.wb.sheetnames:
            print(f'当前表格不含{title}页')
            return False
        ws = self.wb[title]
        rules = []
        if len(skip) == 1: skip.append(len(list(ws.values)))
        if skip[0] < header_line+1: skip[0] = header_line+1
        for i, row in enumerate(ws.values):
            if i == header_line-1:
                headers = row
                v = headers.index(header)
                if header not in headers:
                    print(f'当前表格不含{header}属性列')
                    return None
            elif skip[0] - 1 <= i <= skip[1] - 1:
                rule = {}
                rule[header] = row[v]
                rules.append(rule)
        return rules

    def read_cell(self, title, address='A1'):
        """
        :param title: 表名
        :param address: excel里的坐标
        :return: 该位置的值
        """
        if title not in self.wb.sheetnames:
            print(f'当前表格不含{title}页')
            return False
        ws = self.wb[title]
        return ws[address].value


if __name__ == '__main__':
    file_path = './goods.xlsx'

    nb = XlWriter(file_path, 'goods')
    # nb.append_row(['h1', 'h2'])
    # nb.append_row(['A2', 'B2'])
    # nb.new_sheet('title2', ['A1', 'B1'])
    # nb.append_row(['A2', 'B2'])
    # nb.select_sheet('title1')
    # nb.append_row(['A3', 'B3'])
    # nb.edit_cell('B2','B200')
    # nb.save()


    reader = XlReader(file_path)
    # print(reader.read('goods', [2, 4]))
    # for i in reader.read('goods'):
    #     print(i['Price'])
    #     print(type(i['Price']))
    # print(reader.read_by_header('goods', 'URL'))
    # print(reader.read_cell('title1', 'A1'))
    # for i in reader.read('goods'):
    #     print(i)
